# Chapter 13: Working with Excel Spreadsheets (NEED MORE PRATICES)

import openpyxl
from openpyxl.styles import Font
import openpyxl.worksheet

# 1. Opening Excel Files
def open_excel_file(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        print(f"Workbook '{file_path}' opened successfully.")
        return workbook
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return None

# 2. Accessing Sheets from the Workbook
def get_sheet(workbook, sheet_name):
    try:
        sheet = workbook[str(sheet_name)]
        print(f"Accessed sheet: {sheet.title}")
        return sheet
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found.")
        return None

# 3. Getting Cells from the Sheet
def get_cell_value(sheet, cell_address):
    try:
        cell = sheet[cell_address]
        print(f"Cell {cell_address} value: {cell.value}")
        return cell.value
    except Exception as e:
        print(f"Error accessing cell {cell_address}: {e}")
        return None

# 4. Accessing a Cell's Value Using the `cell()` Method
def get_cell_by_position(sheet, row, column):
    try:
        cell = sheet.cell(row=row, column=column)
        print(f"Cell ({row}, {column}) value: {cell.value}")
        return cell.value
    except Exception as e:
        print(f"Error accessing cell ({row}, {column}): {e}")
        return None

# 5. Creating and Saving Excel Files
def create_and_save_excel(file_path):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'MySheet'
        
        # Adding data to cells
        sheet['A1'] = 'Hello, world!'
        sheet['B1'] = 42
        sheet['C1'] = '=SUM(B1:B10)'

        workbook.save(file_path)
        print(f"Workbook saved as '{file_path}'.")
    except Exception as e:
        print(f"Error creating or saving workbook: {e}")

# 6. Setting Cell Font Style
def set_cell_font(sheet, cell_address, bold=False, italic=False):
    try:
        font_style = Font(bold=bold, italic=italic)
        sheet[cell_address].font = font_style
        print(f"Set font for cell {cell_address} - Bold: {bold}, Italic: {italic}")
    except Exception as e:
        print(f"Error setting font for cell {cell_address}: {e}")

# 7. Adding and Removing Sheets
def add_and_remove_sheet(workbook, new_sheet_name, sheet_to_remove_name):
    try:
        workbook.create_sheet(title=new_sheet_name)
        print(f"Sheet '{new_sheet_name}' added. Current sheets: {workbook.sheetnames}")

        # Removing a sheet
        if sheet_to_remove_name in workbook.sheetnames:
            sheet_to_remove = workbook[sheet_to_remove_name]
            workbook.remove(sheet_to_remove)
            print(f"Sheet '{sheet_to_remove_name}' removed. Current sheets: {workbook.sheetnames}")
        else:
            print(f"Sheet '{sheet_to_remove_name}' not found. No sheet removed.")
    except Exception as e:
        print(f"Error adding or removing sheet: {e}")

# 8. Reading Data from a Sheet into a List of Lists
def read_sheet_to_list(sheet):
    try:
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        print(f"Data read from sheet: {data}")
        return data
    except Exception as e:
        print(f"Error reading sheet data: {e}")
        return None
